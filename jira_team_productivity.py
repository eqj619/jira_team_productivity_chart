import jiralibrary as jlib
import json
import pandas as pd
import datetime
import openpyxl

def last_day_of_month(any_day):
    next_month = any_day.replace(day=28) + datetime.timedelta(days=4)
    return next_month - datetime.timedelta(days=next_month.day)

def createDaysList(daysList, monthList):
    import time
    from datetime import date

    today = date.today()
    for y in range(today.year-2, today.year):
        for m in range(today.month, today.month+12):
            mm = m%12
            if mm == 0:
                mm = 12
                y += 1
            firstday = datetime.date(y, mm, 1)
            lastday = last_day_of_month(firstday)
            daysList.append([firstday.strftime("%Y-%m-%d"), lastday.strftime("%Y-%m-%d")])
            monthList.append(firstday.strftime("%Y-%m"))

# main
import time
from datetime import date
import sys

try:
    wb = openpyxl.load_workbook("jira_team_productivity_template.xlsx")
    ws = wb.active
except:
    print("error: miss jira_team_productivity_template.xlsx in this directory")
    exit(0)

daysList = []
monthList = []
createDaysList(daysList, monthList)

try:
    args = sys.argv
    json_open = open(args[1], 'r')
    json_load = json.load(json_open)
except:
    print("error: miss json file")
    print("usage: python3 jira_team_productivity.py <json file>")
    wb.close()
    exit(0)

jf = jlib.DoFilter(json_load['jiraAccess'])

projectName = json_load['input']['project']
i = 0
for day in daysList:
    f = 'project in (' + projectName + ') AND issuetype in (Epic, Story, Task, Sub-task) AND created >= ' \
        + day[0] +  ' AND created <= ' + day[1]
    ws.cell(row = i+2, column = 2).value = jf.GetNumOfJiraFilter(f)
    print(" >>> ", f)

    f = 'project in (' + projectName + ') AND issuetype in (Epic, Story, Task, Sub-task) AND resolved >= ' \
        + day[0] +  ' AND resolved <= ' + day[1]
    ws.cell(row = i+2, column = 4).value = jf.GetNumOfJiraFilter(f)
    print(" >>> ", f)
    i += 1

i = 0
for ym in monthList:
    ws.cell(row = i+2, column = 1).value = ym
    i += 1

wb.save("jira_" + projectName.replace(",","").replace(" ", "-") + "-team_performance-" + date.today().strftime("%Y-%m-%d") + ".xlsx")
wb.close()
